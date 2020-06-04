#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon May 25 08:22:01 2020

@author: malkusch
"""

import numpy as np
import cv2 as cv
import os
import re
from openpyxl import load_workbook

class RawDataImporter(object):
    """RawDataImporter
    
    Imports Lasca raw data from exported excel files.
    Transfroms the raw data into gray scale images.
    These images can be used for further analysis.
    
     Attributes:
         rawData: A numpy array. Contains the raw data.
    """
    
    def __init__(self):
        """initializes an instance of the RawDataImporter class"""
        self._frameNumber = int()
        self._dx = int()
        self._dy = int()
        self._rawData = np.zeros([580,752,3], dtype = np.uint16)
        
    def __str__(self):
        """prints informtions about the instance variables of the RawData class
        
        Returns:
            Compound string comprising the class' instance variables'
        """
        string = str("\nRawData information\nframeNumber: %i\ndx: %i\ndy: %i"%
                     (self.frameNumber,
                      self.dx,
                      self.dy)
                     )
        return(string)
               
    def __del__(self):
        """removes an instance of RawDataImporter from the heap."""
        string = str("instance of RawDataImporter removed from heap.")
        print(string)
        
    @property
    def rawData(self) -> np.ndarray:
        """returns the instance variable rawData
        
        Returns:
            A numpy array.
        """
        return(self._rawData)
    
    @property
    def frameNumber(self) -> int:
        return(self._frameNumber)
    
    @property
    def dx(self) -> int:
        return(self._dx)
   
    @property
    def dy(self) -> int:
        return(self._dy)
        
    def fileExists(self, fileName: str):
        return os.path.isfile(fileName)
    
    def sheetExists(self, fileName: str, sheetName: str):
        print("under constr.")
        return(True)
    
    def dataExists(self, fileName:str, sheetName: str):
        return(True)
    
    def readFrameNumber(self, fileName: str, sheetName: str):
        if(self.dataExists(fileName = fileName, sheetName = sheetName)):
            wb = load_workbook(filename = fileName,
                               read_only = True,
                               data_only = True)
            ws = wb[sheetName]
            frameNumberCandidates = re.findall(r'\d+', str(ws['A33'].value))
            if (len(frameNumberCandidates) != 1):
                raise ValueError("Could not determine a unique frame number")
            else:
                self._frameNumber = int(frameNumberCandidates[0])
                
    def readResolution(self, fileName: str, sheetName:str):
        if(self.dataExists(fileName = fileName, sheetName = sheetName)):
            wb = load_workbook(filename = fileName,
                               read_only = True,
                               data_only = True)
            ws = wb[sheetName]
            dimensionCandidates = re.findall(r'\d+', str(ws['A30'].value))
            if (len(dimensionCandidates) != 2):
                raise ValueError("Could not determine a unique Resolution")
            else:
                self._dx = int(dimensionCandidates[0])
                self._dy = int(dimensionCandidates[1])
                
    def readMetadata(self, fileName: str, sheetName: str):
        try:
            self.readFrameNumber(fileName, sheetName)
            self.readResolution(fileName, sheetName)
        except ValueError as error:
            print("error during metadata reading procedure:\n%s" %(error))
            self._dx = 0
            self._dy = 0
            self._frameNumber = 0
        self.updateRawData()
            
    def updateRawData(self):
        self._rawData = np.zeros([self.dy, self.dx, self.frameNumber], dtype = np.uint16)
    
    def readFrame(self, fileName: str, sheetName: str, frameNumber: int):
        """imports Raw data from an excel worksheet
        
        Attributes:
            fileName: A string. Path to the excel file.
            sheetName: A string. Name of the desired worksheet.
            frameNumber: An integer. Nmber of time frame to import.
            
        Raises:
            ValueError: An instance of the ValueError.
        """
        
        wb = load_workbook(filename = fileName,
                           read_only = True,
                           data_only = True)
        ws = wb[sheetName]
        firstRow = 38 + 2 * (frameNumber -1) + 580 * (frameNumber-1)
        
        firstCol = 2
        nCols = 752
        nRows = 580
        allCells = np.array([[cell.value for cell in row] for row in ws.iter_rows()])
        if (((firstRow + nRows - 1) <= np.shape(allCells)[0]) & ((firstCol + nCols - 1) <= np.shape(allCells)[1])):
            self._rawData[:,:,frameNumber-1] = allCells[(firstRow-1):(firstRow-1+nRows), (firstCol-1):(firstCol-1+nCols)].astype(np.uint16)
        else:
            self._rawData[:,:,frameNumber-1] = np.zeros([580,752], dtype = np.uint16)
            raise ValueError("Worksheet %s of file %s misses frame %i."%
                             (sheetName,
                              fileName,
                              frameNumber)
                             )
            
    def readData(self, fileName: str, sheetName: str):
        for i in np.arange(0, self.frameNumber, 1):
            self.readFrame(fileName, sheetName, frameNumber = i+1)
            
    def frame(self, frame: int):
        if (self.frameNumber > frame):
            return(self._rawData[:,:,frame])
        else:
            return(None)
        
    def writeFrameToImage(self, frame: int, fileName: str):
        if (self.frameNumber > frame):
            cv.imwrite(fileName, self.rawData[:,:,frame])
            print("Wrote raw channel %i to %s" %(frame, fileName))


def main_loop():
    data = RawDataImporter()
    testPersons = np.arange(1,20,1)
    for person in testPersons:
        idx = str(person).zfill(2)

        fileName =str("/Users/malkusch/PowerFolders/LaSCA/rawData/Fluxauswertung SCTCAPS %s.xlsm" % (idx))
        wb = load_workbook(filename = fileName,
                           read_only = True,
                           data_only = True)
        sheetNames = wb.sheetnames
        for sheetName in sheetNames:
            data.readMetadata(fileName, sheetName)
            data.readData(fileName, sheetName)
            for i in np.arange(0, data.frameNumber, 1):
                outFileName = str("/Users/malkusch/PowerFolders/LaSCA/rawData/SCTCAPS%s/SCTCAPS%s_%s_frame%i.tiff"%
                              (idx,
                               idx,
                               sheetName,
                               i))
                data.writeFrameToImage(frame = i, fileName = outFileName)
    print("done!")
    
def main():
    person = 1
    idx = str(person).zfill(2)
    data = RawDataImporter()
    fileName =str("/Users/malkusch/PowerFolders/LaSCA/rawData/Fluxauswertung SCTCAPS %s.xlsm" % (idx))
    sheetName = str("1.15")
    data.readMetadata(fileName, sheetName)
    data.readData(fileName, sheetName)
    i=0
    outFileName = str("/Users/malkusch/PowerFolders/LaSCA/rawData/SCTCAPS%s/SCTCAPS%s_%s_frame%i.tiff"%
                      (idx, idx, sheetName, i))
    data.writeFrameToImage(frame = i, fileName = outFileName)
    
    
if __name__ == '__main__':
    import doctest
    doctest.testmod(verbose=True)
    main_loop()
        
