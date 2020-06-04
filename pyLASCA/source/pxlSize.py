#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 28 10:54:03 2020

@author: malkusch
"""

import numpy as np
import pandas as pd
import os
import re
from openpyxl import load_workbook

class PxlSize(object):
    def __init__(self):
        self._nPxlX = int()
        self._nPxlY = int()
        self._sizeX = float()
        self._sizeY = float()
        self._pxlSize = float()
        
    def __str__(self):
        message = str("Instance Variables of PxlSize class:\nnumber of pxls i x: %i\nnumberof pxls in y: %i\nwidth in mm: %.3f\n height in mm: %.3f\npxl size in mm: %.3f"%
                      (self.nPxlX, self.nPxlY, self.sizeX, self.sizeY, self.pxlSize))
        return(message)
    
    def __del__(self):
        message = str("Removed instance of class PxlSize from heap.")
        print(message)
        
    @property
    def nPxlX(self) -> int:
        return(self._nPxlX)
    
    @property
    def nPxlY(self) -> int:
        return(self._nPxlY)
    
    @property
    def sizeX(self) -> float:
        return(self._sizeX)
    
    @property
    def sizeY(self) -> float:
        return(self._sizeY)
    
    @property
    def pxlSize(self) -> float:
        return(self._pxlSize)
        
    def importFormExcel(self, fileName: str, sheetName: str):
        if not (os.path.isfile(fileName)):
            raise ValueError("The fileName %s seems to be incorrect" %(fileName))
        wb = load_workbook(filename = fileName,
                               read_only = True,
                               data_only = True)
        ws = wb[sheetName]
        numberOfPixels = re.findall(r'\d+', str(ws['A30'].value))
        if (len(numberOfPixels) != 2):
            raise ValueError("Wrong number of pixels. Got %i needs to be 2" %(len(numberOfPixels)))
        
        self._nPxlX = int(numberOfPixels[0])
        self._nPxlY = int(numberOfPixels[1])
        
        sizeOfFrame = re.findall(r'\d+\,\d+', str(ws['B32'].value))
        if (len(sizeOfFrame) != 2):
            raise ValueError("Wrong number of pixels. Got %i needs to be 2" %(len(sizeOfFrame)))
                             
        self._sizeX = float(sizeOfFrame[0].replace(",", "."))
        self._sizeY = float(sizeOfFrame[1].replace(",", "."))
        
        self.calculatePxlSize()
        
    def calculatePxlSize(self):
       self._pxlSize = 0.5 * ((self.sizeX / self.nPxlX) + (self.sizeY / self.nPxlY)) 

        
        
def main():
    df = pd.DataFrame({"object_id": [0],
                       "measurement": ["none"], 
                       "pxl_X": [0],
                       "pxl_Y": [0],
                       "width_mm": [0.0],
                       "height_mm": [0.0],
                       "pxlSize_mm": [0.0]})

    ps = PxlSize()
    persons = np.arange(1,20,1)
    for person in persons:
        objectId = str(person).zfill(2)
        fileName =str("/Users/malkusch/PowerFolders/LaSCA/rawData/Fluxauswertung SCTCAPS %s.xlsm" %(objectId))
        wb = load_workbook(filename = fileName,
                           read_only = True,
                           data_only = True)
        sheetNames = wb.sheetnames
        for sheetName in sheetNames:
            ps.importFormExcel(fileName = fileName, sheetName = sheetName)
            df_temp = pd.DataFrame({"object_id": [person],
                                    "measurement": [sheetName], 
                                    "pxl_X": [ps.nPxlX],
                                    "pxl_Y": [ps.nPxlY],
                                    "width_mm": [ps.sizeX],
                                    "height_mm": [ps.sizeY],
                                    "pxlSize_mm": [ps.pxlSize]})
            df = df.append(df_temp, ignore_index = True)
    df.drop(df.index[:1], inplace=True)
    df.to_csv(path_or_buf = "/Users/malkusch/PowerFolders/LaSCA/Probanden Arm Bilder mit Markierung/pxl_size.csv")
    print("done")
    
if __name__ == '__main__':
    main()